#!/usr/bin/env python
"""
extend_gmac_3dB.py — Extend GMAC 3dB simulations with more codewords.

Uses MC designs from designs/ directory and decode_single (non-vectorized)
which auto-dispatches to the fastest decoder per path type.
"""

import os
import sys
import time
import numpy as np
from datetime import datetime

sys.path.insert(0, os.path.join(os.path.dirname(__file__), ".."))

from polar.design_mc import load_design, _select_info_frozen
from polar.encoder import polar_encode, build_message
from polar.channels import GaussianMAC
from polar.design import make_path
from polar.decoder import decode_single
from polar.decoder_scl import decode_single_list

import openpyxl

DESIGNS_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "..", "designs")
RESULTS_BASE = os.path.join(os.path.dirname(os.path.abspath(__file__)),
                            "..", "results", "gmac_snr3dB")


def log(msg=""):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def run_sim(code_class, L, N, ku, kv, ncw, seed, channel):
    """Run simulation for one (class, L, N, ku, kv) config."""
    n = int(np.log2(N))
    pfrac = {"A": 0.375, "B": 0.5, "C": 1.0}[code_class]
    path_i = round(pfrac * N)

    # Load MC design
    design_file = os.path.join(DESIGNS_DIR,
                               f"gmac_{code_class}_n{n}_snr3dB.npz")
    su, sv, _, _, pi = load_design(design_file)
    Au, fu = _select_info_frozen(N, su, ku)
    Av, fv = _select_info_frozen(N, sv, kv)
    b = make_path(N, path_i=path_i)

    rng = np.random.default_rng(seed)
    errors = 0
    t0 = time.time()

    for trial in range(ncw):
        info_u = rng.integers(0, 2, ku).tolist()
        info_v = rng.integers(0, 2, kv).tolist()
        u = build_message(N, info_u, Au)
        v = build_message(N, info_v, Av)
        x = polar_encode(u.tolist())
        y = polar_encode(v.tolist())
        z = channel.sample_batch(np.array(x), np.array(y)).tolist()

        if L == 1:
            u_dec, v_dec = decode_single(N, z, b, fu, fv, channel,
                                         log_domain=True)
        else:
            u_dec, v_dec = decode_single_list(N, z, b, fu, fv, channel,
                                              log_domain=True, L=L)

        if not (all(u_dec[p-1] == bit for p, bit in zip(Au, info_u)) and
                all(v_dec[p-1] == bit for p, bit in zip(Av, info_v))):
            errors += 1

        if (trial + 1) % 2000 == 0:
            elapsed = time.time() - t0
            log(f"    {trial+1}/{ncw} done, errors={errors}, "
                f"{elapsed:.1f}s")

    elapsed = time.time() - t0
    return errors, elapsed


def extend_config(code_class, L, ku_kv_map, existing_data, target_cw_map,
                  channel, xlsx_path, title, subtitle, capacity_line):
    """Extend a simulation config with more codewords."""
    N_LIST = [8, 16, 32, 64, 128, 256, 512, 1024]
    results = {}

    for N in N_LIST:
        ku, kv = ku_kv_map[N]
        ex_cw = existing_data[N]["cw"]
        ex_err = existing_data[N]["err"]
        target = target_cw_map.get(N, ex_cw)
        new_needed = max(0, target - ex_cw)

        if new_needed <= 0:
            log(f"  N={N:>5d}: already have {ex_cw} CW -- SKIP")
            results[N] = {"err": ex_err, "cw": ex_cw, "ku": ku, "kv": kv}
            continue

        log(f"  N={N:>5d}: running {new_needed} new CW "
            f"(have {ex_cw}, target {target})")

        # Use different seed to avoid correlation with existing results
        seed = 777777 + N * 1000 + L * 100

        new_err, elapsed = run_sim(code_class, L, N, ku, kv,
                                   new_needed, seed, channel)

        total_err = ex_err + new_err
        total_cw = ex_cw + new_needed
        bler = total_err / total_cw
        log(f"  N={N:>5d}: BLER={bler:.5f} ({total_err}/{total_cw}) "
            f"in {elapsed:.1f}s")
        results[N] = {"err": total_err, "cw": total_cw, "ku": ku, "kv": kv}

    # Write xlsx
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([title] + [None] * 7)
    ws.append([subtitle] + [None] * 7)
    ws.append([None] * 8)
    ws.append(['N', 'ku', 'kv', 'Ru', 'Rv', 'Sum Rate', 'BLER', 'Codewords'])
    for N in N_LIST:
        r = results[N]
        ku, kv = r["ku"], r["kv"]
        Ru, Rv = ku / N, kv / N
        bler = r["err"] / r["cw"] if r["cw"] > 0 else 0
        ws.append([N, ku, kv, round(Ru, 4), round(Rv, 4),
                   round(Ru + Rv, 4), round(bler, 6), r["cw"]])
    ws.append([None] * 8)
    ws.append([capacity_line] + [None] * 7)
    wb.save(xlsx_path)
    log(f"  Saved: {xlsx_path}")
    return results


def read_existing_xlsx(path):
    """Read existing xlsx results."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=5, max_row=12, values_only=True):
        if row[0] and isinstance(row[0], (int, float)) and row[0] > 1:
            N = int(row[0])
            ku, kv = int(row[1]), int(row[2])
            cw = int(row[7])
            bler = float(row[6])
            data[N] = {
                "ku": ku, "kv": kv, "cw": cw,
                "err": round(bler * cw)
            }
    return data


def main():
    sigma2 = 10.0 ** (-3.0 / 10.0)
    channel = GaussianMAC(sigma2=sigma2)
    I_ZX, I_ZY_X, I_ZXY = channel.capacity()
    cap_line = f"Capacity: I(Z;X)={I_ZX:.4f}, I(Z;Y|X)={I_ZY_X:.4f}, I(Z;X,Y)={I_ZXY:.4f}"
    log(f"GMAC 3dB: {cap_line}")

    N_LIST = [8, 16, 32, 64, 128, 256, 512, 1024]

    # =============================================
    # 1. ClassB Ru39/Rv39 SC -> 20K CW
    # =============================================
    log("\n" + "=" * 60)
    log("1. ClassB Ru39/Rv39 SC -- 5K->20K CW")
    log("=" * 60)

    existing = read_existing_xlsx(os.path.join(
        RESULTS_BASE, "gmac_snr3dB_classB_Ru39_Rv39_SC",
        "classB_Ru39_Rv39_snr3dB_SC.xlsx"))
    ku_kv = {N: (existing[N]["ku"], existing[N]["kv"]) for N in N_LIST}
    target = {N: 20000 for N in N_LIST}

    extend_config("B", 1, ku_kv, existing, target, channel,
        os.path.join(RESULTS_BASE, "gmac_snr3dB_classB_Ru39_Rv39_SC",
                     "classB_Ru39_Rv39_snr3dB_SC.xlsx"),
        "Gaussian MAC -- Class B, (Ru=0.39, Rv=0.39), SNR=3dB, SC (L=1)",
        "Symmetric capacity: Rx=Ry <= 0.5548 bits/use | Sum capacity: 1.1095 bits/use | Path: 0^(N/2) 1^N 0^(N/2)",
        cap_line)

    # =============================================
    # 2. ClassB Ru28/Rv28 SC -> N=1024 to 20K
    # =============================================
    log("\n" + "=" * 60)
    log("2. ClassB Ru28/Rv28 SC -- N=1024: 3K->20K CW")
    log("=" * 60)

    existing2 = read_existing_xlsx(os.path.join(
        RESULTS_BASE, "gmac_snr3dB_classB_Ru28_Rv28_SC",
        "classB_Ru28_Rv28_snr3dB_SC.xlsx"))
    ku_kv2 = {N: (existing2[N]["ku"], existing2[N]["kv"]) for N in N_LIST}
    target2 = {1024: 20000}  # Only extend N=1024

    extend_config("B", 1, ku_kv2, existing2, target2, channel,
        os.path.join(RESULTS_BASE, "gmac_snr3dB_classB_Ru28_Rv28_SC",
                     "classB_Ru28_Rv28_snr3dB_SC.xlsx"),
        "Gaussian MAC -- Class B, (Ru=0.28, Rv=0.28), SNR=3dB, SC (L=1)",
        "Symmetric capacity: Rx=Ry <= 0.5548 bits/use | Sum capacity: 1.1095 bits/use | Path: 0^(N/2) 1^N 0^(N/2)",
        cap_line)

    # =============================================
    # 3. ClassC Ru19/Rv36 SC -> 10K CW
    # =============================================
    log("\n" + "=" * 60)
    log("3. ClassC Ru19/Rv36 SC -- 2K->10K CW")
    log("=" * 60)

    existing3 = read_existing_xlsx(os.path.join(
        RESULTS_BASE, "gmac_snr3dB_classC_Ru19_Rv36_SC",
        "classC_Ru19_Rv36_snr3dB_SC.xlsx"))
    ku_kv3 = {N: (existing3[N]["ku"], existing3[N]["kv"]) for N in N_LIST}
    target3 = {N: 10000 for N in N_LIST}

    extend_config("C", 1, ku_kv3, existing3, target3, channel,
        os.path.join(RESULTS_BASE, "gmac_snr3dB_classC_Ru19_Rv36_SC",
                     "classC_Ru19_Rv36_snr3dB_SC.xlsx"),
        "Gaussian MAC -- Class C, (Ru=0.19, Rv=0.36), SNR=3dB, SC (L=1)",
        "Symmetric capacity: Rx=Ry <= 0.5548 bits/use | Sum capacity: 1.1095 bits/use | Path: 0^N 1^N",
        cap_line)

    # =============================================
    # 4. ClassA Ru36/Rv19 SC -> N=1024 to 10K
    # =============================================
    log("\n" + "=" * 60)
    log("4. ClassA Ru36/Rv19 SC -- N=1024: 3K->10K CW")
    log("=" * 60)

    existing4 = read_existing_xlsx(os.path.join(
        RESULTS_BASE, "gmac_snr3dB_classA_Ru36_Rv19_SC",
        "classA_Ru36_Rv19_snr3dB_SC.xlsx"))
    ku_kv4 = {N: (existing4[N]["ku"], existing4[N]["kv"]) for N in N_LIST}
    target4 = {1024: 10000}

    extend_config("A", 1, ku_kv4, existing4, target4, channel,
        os.path.join(RESULTS_BASE, "gmac_snr3dB_classA_Ru36_Rv19_SC",
                     "classA_Ru36_Rv19_snr3dB_SC.xlsx"),
        "Gaussian MAC -- Class A, (Ru=0.36, Rv=0.19), SNR=3dB, SC (L=1)",
        "Symmetric capacity: Rx=Ry <= 0.5548 bits/use | Sum capacity: 1.1095 bits/use | Path: 0^N 1^N",
        cap_line)

    # =============================================
    # 5. ClassB Ru39/Rv39 SCL-32
    # =============================================
    log("\n" + "=" * 60)
    log("5. ClassB Ru39/Rv39 SCL-32 -- extending CW")
    log("=" * 60)

    existing5 = read_existing_xlsx(os.path.join(
        RESULTS_BASE, "gmac_snr3dB_classB_Ru39_Rv39_SCL32",
        "classB_Ru39_Rv39_snr3dB_SCL_L32.xlsx"))
    ku_kv5 = {N: (existing5[N]["ku"], existing5[N]["kv"]) for N in N_LIST}
    # 5K for small N, 3K for N=512, 2K for N=1024
    target5 = {8: 5000, 16: 5000, 32: 5000, 64: 5000,
               128: 5000, 256: 5000, 512: 3000, 1024: 2000}

    extend_config("B", 32, ku_kv5, existing5, target5, channel,
        os.path.join(RESULTS_BASE, "gmac_snr3dB_classB_Ru39_Rv39_SCL32",
                     "classB_Ru39_Rv39_snr3dB_SCL_L32.xlsx"),
        "Gaussian MAC -- Class B, (Ru=0.39, Rv=0.39), SNR=3dB, SCL (L=32)",
        "Symmetric capacity: Rx=Ry <= 0.5548 bits/use | Sum capacity: 1.1095 bits/use | Path: 0^(N/2) 1^N 0^(N/2)",
        cap_line)

    # =============================================
    # 6. ClassC Ru19/Rv36 SCL-32
    # =============================================
    log("\n" + "=" * 60)
    log("6. ClassC Ru19/Rv36 SCL-32 -- extending CW")
    log("=" * 60)

    existing6 = read_existing_xlsx(os.path.join(
        RESULTS_BASE, "gmac_snr3dB_classC_Ru19_Rv36_SCL32",
        "classC_Ru19_Rv36_snr3dB_SCL_L32.xlsx"))
    ku_kv6 = {N: (existing6[N]["ku"], existing6[N]["kv"]) for N in N_LIST}
    # 3K for N<=128, 2K for N=256,512, 1K for N=1024
    target6 = {8: 3000, 16: 3000, 32: 3000, 64: 3000,
               128: 3000, 256: 2000, 512: 2000, 1024: 1000}

    extend_config("C", 32, ku_kv6, existing6, target6, channel,
        os.path.join(RESULTS_BASE, "gmac_snr3dB_classC_Ru19_Rv36_SCL32",
                     "classC_Ru19_Rv36_snr3dB_SCL_L32.xlsx"),
        "Gaussian MAC -- Class C, (Ru=0.19, Rv=0.36), SNR=3dB, SCL (L=32)",
        "Symmetric capacity: Rx=Ry <= 0.5548 bits/use | Sum capacity: 1.1095 bits/use | Path: 0^N 1^N",
        cap_line)

    log("\n" + "=" * 60)
    log("ALL DONE!")
    log("=" * 60)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("\n*** Interrupted ***")
        sys.exit(1)

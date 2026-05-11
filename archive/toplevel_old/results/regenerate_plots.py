#!/usr/bin/env python3
"""Regenerate all BLER plot PNGs with correct y-axis limits."""

import os
import sys
import json
import math
import numpy as np
import openpyxl

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.ticker import ScalarFormatter

RESULTS_DIR = '/Users/ytnspybq/PycharmProjects/polar_codes_MAC/to_git_v2/results'


def read_excel(xlsx_path):
    """Read Excel file, return title, subtitle, headers, data rows."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    title = ws.cell(1, 1).value or ''
    subtitle = ws.cell(2, 1).value or ''

    # Row 4 = headers
    max_col = ws.max_column
    headers = []
    for c in range(1, max_col + 1):
        v = ws.cell(4, c).value
        if v is not None:
            headers.append(str(v))
        else:
            break

    # Row 5+ = data (until None in col 1)
    data = []
    for r in range(5, ws.max_row + 1):
        n_val = ws.cell(r, 1).value
        if n_val is None or (isinstance(n_val, str) and not n_val.replace('.','').isdigit()):
            break
        row = []
        for c in range(1, len(headers) + 1):
            row.append(ws.cell(r, c).value)
        data.append(row)

    return title, subtitle, headers, data


def detect_plot_type(headers):
    """Detect: '3decoder', '2decoder', 'single'."""
    h_set = set(headers)
    if 'NN BLER' in h_set and 'SC BLER' in h_set and 'SCL BLER' in h_set:
        return '3decoder'
    elif 'NN BLER' in h_set and 'SC BLER' in h_set:
        return '2decoder'
    elif 'BLER' in h_set:
        return 'single'
    else:
        print(f"  WARNING: Unknown header set: {headers}")
        return None


def get_col(headers, name):
    """Get column index for a header name."""
    for i, h in enumerate(headers):
        if h == name:
            return i
    return None


def compute_ylim_bottom(all_bler_values):
    """Compute ylim bottom: half a decade below min non-zero BLER."""
    nonzero = [b for b in all_bler_values if b is not None and b > 0]
    if not nonzero:
        return 1e-3  # fallback
    min_bler = min(nonzero)
    ylim_bottom = 10 ** (math.floor(math.log10(min_bler)) - 0.5)
    return ylim_bottom


def split_title(title, max_len=60):
    """Split title into 2 lines at a natural break if > max_len chars."""
    if len(title) <= max_len:
        return title
    # Try splitting at comma nearest to middle
    mid = len(title) // 2
    # Find commas
    commas = [i for i, c in enumerate(title) if c == ',']
    if commas:
        best = min(commas, key=lambda x: abs(x - mid))
        return title[:best+1].strip() + '\n' + title[best+1:].strip()
    # Try splitting at space nearest to middle
    spaces = [i for i, c in enumerate(title) if c == ' ']
    if spaces:
        best = min(spaces, key=lambda x: abs(x - mid))
        return title[:best].strip() + '\n' + title[best+1:].strip()
    return title


def make_plot(xlsx_path, png_path):
    """Generate one plot from xlsx, save to png."""
    title, subtitle, headers, data = read_excel(xlsx_path)
    plot_type = detect_plot_type(headers)
    if plot_type is None:
        print(f"  SKIPPED (unknown type): {xlsx_path}")
        return

    n_col = get_col(headers, 'N')
    ns = [row[n_col] for row in data]
    n_values = [int(n) for n in ns]

    # Gather all BLER columns
    bler_cols = {}
    if plot_type == '3decoder':
        bler_cols['NN BLER'] = get_col(headers, 'NN BLER')
        bler_cols['SC BLER'] = get_col(headers, 'SC BLER')
        bler_cols['SCL BLER'] = get_col(headers, 'SCL BLER')
    elif plot_type == '2decoder':
        bler_cols['NN BLER'] = get_col(headers, 'NN BLER')
        bler_cols['SC BLER'] = get_col(headers, 'SC BLER')
    else:  # single
        bler_cols['BLER'] = get_col(headers, 'BLER')

    # Collect all non-zero BLERs for ylim computation
    all_bler = []
    for row in data:
        for col_idx in bler_cols.values():
            v = row[col_idx]
            if v is not None and v > 0:
                all_bler.append(v)

    ylim_bottom = compute_ylim_bottom(all_bler)

    # Get sum rate column
    sum_rate_col = get_col(headers, 'Sum Rate')
    # Get Ru, Rv columns
    ru_col = get_col(headers, 'Ru')
    rv_col = get_col(headers, 'Rv')
    # Get NN/SC ratio column
    ratio_col = get_col(headers, 'NN/SC')

    # ---- Create figure ----
    fig, ax = plt.subplots(figsize=(9, 5.5), dpi=150)

    is_nn_plot = plot_type in ('2decoder', '3decoder')

    if is_nn_plot:
        x_values = np.array(n_values, dtype=float)
    else:
        x_values = np.arange(len(n_values))

    # Plot each BLER series
    color_map = {
        'NN BLER': '#2166ac',
        'SC BLER': '#b2182b',
        'SCL BLER': '#2ca02c',
        'BLER': '#2166ac',
    }
    marker_map = {
        'NN BLER': 'o',
        'SC BLER': 's',
        'SCL BLER': '^',
        'BLER': 'o',
    }
    style_map = {
        'NN BLER': '-',
        'SC BLER': '--',
        'SCL BLER': '-',
        'BLER': '-',
    }
    label_map = {
        'NN BLER': 'Neural SC',
        'SC BLER': 'SC',
        'SCL BLER': 'SCL (L=32)',
        'BLER': None,  # will be set from title
    }

    for bler_name, col_idx in bler_cols.items():
        color = color_map[bler_name]
        marker = marker_map[bler_name]
        linestyle = style_map[bler_name]
        label = label_map[bler_name]

        if bler_name == 'BLER':
            # Determine label from title
            if 'SCL' in title:
                label = 'SCL (L=32)'
            else:
                label = 'SC'

        blers = [row[col_idx] for row in data]

        # Separate non-zero and zero points
        x_nonzero, y_nonzero = [], []
        x_zero, y_zero = [], []
        for i, b in enumerate(blers):
            if b is None or b == 0:
                x_zero.append(x_values[i])
                y_zero.append(ylim_bottom)
            else:
                x_nonzero.append(x_values[i])
                y_nonzero.append(b)

        if x_nonzero:
            ax.semilogy(x_nonzero, y_nonzero, marker=marker, linestyle=linestyle,
                        color=color, lw=2, ms=8, label=label, zorder=3)
        if x_zero:
            zero_label = f'{label} (BLER=0)' if label else 'BLER=0'
            ax.semilogy(x_zero, y_zero, marker='v', linestyle='none',
                        color=color, ms=8, label=zero_label, zorder=3)

    # Sum rate annotations on non-zero points
    if sum_rate_col is not None:
        # Annotate on the first BLER series that has data
        first_bler_name = list(bler_cols.keys())[0]
        first_col = bler_cols[first_bler_name]
        for i, row in enumerate(data):
            b = row[first_col]
            sr = row[sum_rate_col]
            if b is not None and b > 0 and sr is not None:
                ax.annotate(f'R={sr:.3f}', (x_values[i], b),
                            textcoords='offset points', xytext=(8, 8),
                            fontsize=7, color='gray', zorder=4)
    elif ru_col is not None and rv_col is not None:
        # For nn_vs_sc gmac files without Sum Rate, annotate Ru+Rv
        first_bler_name = list(bler_cols.keys())[0]
        first_col = bler_cols[first_bler_name]
        for i, row in enumerate(data):
            b = row[first_col]
            ru = row[ru_col]
            rv = row[rv_col]
            if b is not None and b > 0 and ru is not None and rv is not None:
                sr = ru + rv
                ax.annotate(f'R={sr:.3f}', (x_values[i], b),
                            textcoords='offset points', xytext=(8, 8),
                            fontsize=7, color='gray', zorder=4)

    # Ratio annotations for NN plots
    if is_nn_plot and ratio_col is not None:
        nn_col = bler_cols.get('NN BLER')
        if nn_col is not None:
            for i, row in enumerate(data):
                nn_b = row[nn_col]
                ratio = row[ratio_col]
                if nn_b is not None and nn_b > 0 and ratio is not None:
                    color_r = '#2ca02c' if ratio < 1 else 'gray'
                    ax.annotate(f'{ratio:.2f}x', (x_values[i], nn_b),
                                textcoords='offset points', xytext=(0, -14),
                                fontsize=8, color=color_r, ha='center', zorder=4)

    # X-axis
    if is_nn_plot:
        ax.set_xscale('log', base=2)
        ax.xaxis.set_major_formatter(ScalarFormatter())
        ax.set_xticks(n_values)
        ax.get_xaxis().set_tick_params(which='minor', size=0)
        ax.get_xaxis().set_tick_params(which='minor', width=0)
    else:
        ax.set_xticks(x_values)
        ax.set_xticklabels([str(n) for n in n_values],
                           rotation=45 if len(n_values) > 8 else 0)

    ax.set_xlabel('Block Length N', fontsize=11)
    ax.set_ylabel('Block Error Rate (BLER)', fontsize=11)

    # Y-axis limits
    ax.set_ylim(bottom=ylim_bottom, top=1.0)

    # Title
    display_title = split_title(title)
    ax.set_title(display_title, fontsize=12)

    # Grid and legend
    ax.grid(True, which='both', alpha=0.3)
    ax.legend(fontsize=10)

    plt.tight_layout()
    fig.savefig(png_path, dpi=150)
    plt.close(fig)


def find_all_xlsx_dirs():
    """Find all xlsx files (excluding no256), return list of (xlsx_path, png_path) tuples."""
    pairs = []
    for root, dirs, files in os.walk(RESULTS_DIR):
        if 'no256' in root:
            continue
        xlsx_files = [f for f in files if f.endswith('.xlsx') and 'no256' not in f]
        png_files = [f for f in files if f.endswith('.png') and 'no256' not in f]
        for xlsx_f in xlsx_files:
            xlsx_path = os.path.join(root, xlsx_f)
            # Find corresponding PNG
            if png_files:
                png_path = os.path.join(root, png_files[0])
            else:
                # Generate png name from xlsx name
                png_path = os.path.join(root, xlsx_f.replace('.xlsx', '.png'))
            pairs.append((xlsx_path, png_path))
    return pairs


def regenerate_best_pdf():
    """Regenerate best_results.pdf from _best_plots.json."""
    json_path = os.path.join(RESULTS_DIR, '_best_plots.json')
    if not os.path.exists(json_path):
        print("No _best_plots.json found, skipping PDF generation.")
        return

    with open(json_path) as f:
        plot_list = json.load(f)

    # Filter: remove nn_vs_sc if nn_vs_sc_vs_scl exists for same rates
    # Extract rate info from path
    def get_rate_key(path):
        """Extract rate identifier like 'Ru30_Rv60' from path."""
        import re
        m = re.search(r'(Ru\d+_Rv\d+)', path)
        return m.group(1) if m else None

    def get_channel_key(path):
        """Extract channel prefix like 'bemac' or 'gmac_snr6dB'."""
        parts = path.split('/')
        for p in parts:
            if p in ('bemac', 'abnmac', 'gmac_snr3dB', 'gmac_snr6dB'):
                return p
        return None

    # Find all nn_vs_sc_vs_scl entries
    scl_keys = set()
    for p in plot_list:
        if 'nn_vs_sc_vs_scl' in p:
            rk = get_rate_key(p)
            ck = get_channel_key(p)
            if rk and ck:
                scl_keys.add((ck, rk))

    filtered = []
    for p in plot_list:
        if 'nn_vs_sc_vs_scl' not in p and 'nn_vs_sc' in p:
            rk = get_rate_key(p)
            ck = get_channel_key(p)
            if (ck, rk) in scl_keys:
                print(f"  Filtering out (superseded by vs_scl): {os.path.basename(p)}")
                continue
        filtered.append(p)

    # Combine into PDF
    from PIL import Image
    images = []
    for p in filtered:
        if os.path.exists(p):
            img = Image.open(p).convert('RGB')
            images.append(img)
        else:
            print(f"  WARNING: Missing PNG: {p}")

    if images:
        pdf_path = os.path.join(RESULTS_DIR, 'best_results.pdf')
        images[0].save(pdf_path, save_all=True, append_images=images[1:])
        print(f"Saved best_results.pdf with {len(images)} pages")
    else:
        print("No images found for PDF")


def main():
    pairs = find_all_xlsx_dirs()
    print(f"Found {len(pairs)} xlsx/png pairs to regenerate")

    for i, (xlsx_path, png_path) in enumerate(pairs):
        rel = os.path.relpath(xlsx_path, RESULTS_DIR)
        print(f"[{i+1}/{len(pairs)}] {rel}")
        try:
            make_plot(xlsx_path, png_path)
            print(f"  -> {os.path.basename(png_path)} OK")
        except Exception as e:
            print(f"  ERROR: {e}")
            import traceback
            traceback.print_exc()

    print("\n--- Regenerating best_results.pdf ---")
    regenerate_best_pdf()
    print("\nDone!")


if __name__ == '__main__':
    main()

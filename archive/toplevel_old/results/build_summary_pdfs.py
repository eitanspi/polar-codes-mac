#!/usr/bin/env python3
"""
Build two summary PDFs from the curated experiment folders:
  1. all_experiments.pdf  — every experiment, one page each
  2. good_results.pdf     — only the experiments classified as "good"

Each page has:
  - Title (channel, class, rates, decoder)
  - Embedded plot (existing PNG from the experiment folder)
  - Data table from the .xlsx
  - One-line quality comment generated from data + good_weird/ classification
"""

import os
import re
import math
import openpyxl
import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.backends.backend_pdf import PdfPages
from matplotlib import image as mpimg

RESULTS_DIR = "/Users/ytnspybq/PycharmProjects/polar_codes_MAC/to_git_v2/results"
MAIN_DIRS = ["bemac", "abnmac", "gmac_snr3dB", "gmac_snr6dB"]
GOOD_WEIRD_DIR = os.path.join(RESULTS_DIR, "good_weird")

# Curated allowlist: experiments that are presentation-quality even if the
# good_weird/ classification didn't put them in "good" (e.g. headline NN
# results that beat SC, clean baselines worth showing as reference).
PRESENTATION_GOOD = {
    # BEMAC headline neural results (NN beats SC)
    "bemac_classC_Ru30_Rv60_nn_vs_sc",
    "bemac_classC_Ru35_Rv70_nn_vs_sc",
    "bemac_classB_Ru50_Rv70_nn_vs_sc",
    "bemac_classC_Ru35_Rv70_nn_vs_sc_vs_scl",
    # BEMAC neural SCL
    "bemac_classB_Ru50_Rv70_nn_scl",
    # BEMAC clean SC/SCL baselines worth showing as reference
    "bemac_classC_Ru30_Rv60_SC",
    "bemac_classB_Ru50_Rv70_scl32",
    # GMAC clean baselines + the solid NN result
    "gmac_snr6dB_classB_Ru48_Rv48_SCL32",
    "gmac_snr6dB_classB_Ru48_Rv48_nn_scl",
}


# ---------- xlsx parsing ----------

def read_xlsx(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    title = ws.cell(1, 1).value or ""
    subtitle = ws.cell(2, 1).value or ""
    headers = []
    for c in range(1, ws.max_column + 1):
        v = ws.cell(4, c).value
        if v is None:
            break
        headers.append(str(v))
    data = []
    for r in range(5, ws.max_row + 1):
        n_val = ws.cell(r, 1).value
        if n_val is None:
            break
        if isinstance(n_val, str) and not n_val.replace(".", "").isdigit():
            break
        row = [ws.cell(r, c).value for c in range(1, len(headers) + 1)]
        data.append(row)
    return title, subtitle, headers, data


def col_index(headers, name):
    for i, h in enumerate(headers):
        if h == name:
            return i
    return None


# ---------- find experiments ----------

def find_experiment_folders():
    """Return list of (channel, folder_path, xlsx_path, png_path) for all main experiments."""
    out = []
    for channel in MAIN_DIRS:
        chan_dir = os.path.join(RESULTS_DIR, channel)
        if not os.path.isdir(chan_dir):
            continue
        for entry in sorted(os.listdir(chan_dir)):
            sub = os.path.join(chan_dir, entry)
            if not os.path.isdir(sub):
                continue
            xlsxs = [f for f in os.listdir(sub) if f.endswith(".xlsx")]
            pngs = [f for f in os.listdir(sub) if f.endswith(".png")]
            if not xlsxs:
                continue
            xlsx_path = os.path.join(sub, xlsxs[0])
            png_path = os.path.join(sub, pngs[0]) if pngs else None
            out.append((channel, entry, xlsx_path, png_path))
    return out


def build_quality_index():
    """Map experiment folder name → 'good' | 'almost_good' | 'weird' | None."""
    idx = {}
    for label in ["good", "almost_good", "weird"]:
        d = os.path.join(GOOD_WEIRD_DIR, label)
        if not os.path.isdir(d):
            continue
        for entry in os.listdir(d):
            if os.path.isdir(os.path.join(d, entry)):
                idx[entry] = label
    return idx


# ---------- judgement ----------

def classify_columns(headers):
    h = set(headers)
    has_nn = "NN BLER" in h
    has_sc = "SC BLER" in h
    has_scl = "SCL BLER" in h
    has_single = "BLER" in h
    if has_nn and has_sc and has_scl:
        return "nn_sc_scl"
    if has_nn and has_sc:
        return "nn_sc"
    if has_single:
        return "single"
    return "unknown"


def is_monotonic_nondecreasing_then_decreasing(values):
    """Approximately monotonic with small N spikes allowed."""
    nz = [(i, v) for i, v in enumerate(values) if v is not None and v > 0]
    if len(nz) < 3:
        return True
    # Check that the second half is monotonically decreasing
    half = len(nz) // 2
    tail = [v for _, v in nz[half:]]
    for a, b in zip(tail, tail[1:]):
        if b > a * 1.2:  # 20% slack
            return False
    return True


def make_comment(headers, data, quality_label, title):
    """Return a single-line comment about this experiment."""
    kind = classify_columns(headers)

    n_col = col_index(headers, "N")
    cw_cols = [col_index(headers, c) for c in ["NN CW", "SC CW", "SCL CW", "CW", "Codewords"] if col_index(headers, c) is not None]
    all_cw = []
    for row in data:
        for c in cw_cols:
            v = row[c]
            if v is not None and isinstance(v, (int, float)):
                all_cw.append(int(v))
    min_cw = min(all_cw) if all_cw else None

    bits = []

    # CW reliability
    if min_cw is None:
        cw_tag = "unknown CW"
    elif min_cw >= 5000:
        cw_tag = f"{min_cw}+ cw, solid"
    elif min_cw >= 1000:
        cw_tag = f"{min_cw}+ cw, moderate"
    else:
        cw_tag = f"only {min_cw} cw, low confidence"
    bits.append(cw_tag)

    # Quality label from good_weird/
    if quality_label == "good":
        bits.append("classified GOOD")
    elif quality_label == "almost_good":
        bits.append("almost-good")
    elif quality_label == "weird":
        bits.append("flagged WEIRD")

    # NN ratio commentary
    if kind in ("nn_sc", "nn_sc_scl"):
        nn_col = col_index(headers, "NN BLER")
        sc_col = col_index(headers, "SC BLER")
        ratio_col = col_index(headers, "NN/SC")
        nn_vals = [row[nn_col] for row in data]
        sc_vals = [row[sc_col] for row in data]
        ratios = []
        for nv, sv in zip(nn_vals, sc_vals):
            if nv and sv and nv > 0 and sv > 0:
                ratios.append(nv / sv)
        if ratios:
            min_r = min(ratios)
            max_r = max(ratios)
            if max_r <= 1.0:
                bits.append(f"NN beats SC at all N (ratio {min_r:.2f}–{max_r:.2f}x)")
            elif min_r <= 1.0 and max_r <= 1.5:
                bits.append(f"NN matches SC ({min_r:.2f}–{max_r:.2f}x)")
            elif min_r <= 1.5:
                bits.append(f"NN close at small N, diverges large N ({min_r:.2f}–{max_r:.2f}x)")
            else:
                bits.append(f"NN worse than SC ({min_r:.2f}–{max_r:.2f}x)")

    # Monotonic check on the primary BLER column
    primary = None
    if kind == "nn_sc" or kind == "nn_sc_scl":
        primary = col_index(headers, "SC BLER")
    elif kind == "single":
        primary = col_index(headers, "BLER")
    if primary is not None:
        vals = [row[primary] for row in data]
        if not is_monotonic_nondecreasing_then_decreasing(vals):
            bits.append("non-monotonic (likely small-N rate quantization)")

    return " · ".join(bits)


# ---------- page rendering ----------

def split_title(title, max_len=70):
    if len(title) <= max_len:
        return title
    mid = len(title) // 2
    spaces = [i for i, c in enumerate(title) if c == " "]
    if spaces:
        best = min(spaces, key=lambda x: abs(x - mid))
        return title[:best].strip() + "\n" + title[best + 1:].strip()
    return title


def format_cell(val, header):
    if val is None:
        return ""
    if isinstance(val, float):
        if header in ("Ru", "Rv", "Sum Rate", "NN/SC"):
            return f"{val:.3f}"
        if "BLER" in header:
            if val == 0:
                return "0"
            return f"{val:.2e}"
        return f"{val:.4g}"
    return str(val)


def render_page(pdf, channel, name, title, subtitle, headers, data, png_path, comment, quality_label):
    fig = plt.figure(figsize=(8.5, 11), dpi=100)
    fig.patch.set_facecolor("white")

    # 1) title strip
    title_text = split_title(title or name)
    fig.text(0.5, 0.96, title_text, ha="center", va="top", fontsize=12, fontweight="bold")
    if subtitle:
        sub = subtitle if len(subtitle) < 110 else subtitle[:107] + "..."
        fig.text(0.5, 0.925, sub, ha="center", va="top", fontsize=8, color="#444")

    # 2) plot area
    if png_path and os.path.exists(png_path):
        img = mpimg.imread(png_path)
        ax_img = fig.add_axes([0.08, 0.46, 0.84, 0.43])
        ax_img.imshow(img)
        ax_img.axis("off")
    else:
        ax_img = fig.add_axes([0.08, 0.46, 0.84, 0.43])
        ax_img.text(0.5, 0.5, "(no plot)", ha="center", va="center", fontsize=14, color="#888")
        ax_img.axis("off")

    # 3) table
    ax_tbl = fig.add_axes([0.05, 0.12, 0.9, 0.30])
    ax_tbl.axis("off")
    cells = [[format_cell(v, h) for v, h in zip(row, headers)] for row in data]
    if cells:
        tbl = ax_tbl.table(
            cellText=cells,
            colLabels=headers,
            loc="upper center",
            cellLoc="center",
        )
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(7)
        tbl.scale(1, 1.25)
        # color header
        for c in range(len(headers)):
            tbl[(0, c)].set_facecolor("#e0e0e0")
            tbl[(0, c)].set_text_props(weight="bold")

    # 4) comment + quality strip at bottom
    color_map = {"good": "#2ca02c", "almost_good": "#cc7a00", "weird": "#b22222", None: "#444"}
    label_color = color_map.get(quality_label, "#444")
    label_text = quality_label.upper() if quality_label else "UNCATEGORIZED"
    fig.text(0.05, 0.075, f"[{label_text}]", fontsize=9, fontweight="bold", color=label_color)
    fig.text(0.05, 0.05, f"Comment: {comment}", fontsize=8, color="#222", wrap=True)
    fig.text(0.05, 0.025, f"Source: {channel}/{name}", fontsize=6, color="#888")

    pdf.savefig(fig, bbox_inches="tight")
    plt.close(fig)


# ---------- group ordering ----------

def channel_sort_key(channel):
    order = {"bemac": 0, "abnmac": 1, "gmac_snr3dB": 2, "gmac_snr6dB": 3}
    return order.get(channel, 99)


def folder_sort_key(name):
    # group SC < SCL < nn_vs_sc < nn_scl
    if "nn_vs_sc_vs_scl" in name:
        sub = 4
    elif "nn_scl" in name or "neural_scl" in name:
        sub = 3
    elif "nn_vs_sc" in name or "mlp_nn_vs_sc" in name:
        sub = 2
    elif "SCL" in name or "scl" in name:
        sub = 1
    else:
        sub = 0
    return (sub, name)


# ---------- main ----------

def build_pdfs():
    experiments = find_experiment_folders()
    quality_idx = build_quality_index()

    # Sort: by channel, then by class, then SC < SCL < NN
    experiments.sort(key=lambda e: (channel_sort_key(e[0]), folder_sort_key(e[1])))

    print(f"Found {len(experiments)} experiments")
    print(f"Quality index entries: {len(quality_idx)}")

    all_pdf_path = os.path.join(RESULTS_DIR, "all_experiments_summary.pdf")
    good_pdf_path = os.path.join(RESULTS_DIR, "good_results_summary.pdf")

    with PdfPages(all_pdf_path) as all_pdf, PdfPages(good_pdf_path) as good_pdf:
        # ----- cover pages -----
        for pdf, title in [
            (all_pdf, "Polar Codes MAC — All Experiments"),
            (good_pdf, "Polar Codes MAC — Good Results"),
        ]:
            fig = plt.figure(figsize=(8.5, 11), dpi=100)
            fig.text(0.5, 0.6, title, ha="center", fontsize=22, fontweight="bold")
            fig.text(0.5, 0.55, f"{len(experiments)} experiments scanned", ha="center", fontsize=12, color="#444")
            fig.text(0.5, 0.50, "Generated by build_summary_pdfs.py", ha="center", fontsize=9, color="#888")
            fig.text(0.5, 0.45,
                     "Each page: title · plot · data table · auto-generated quality comment",
                     ha="center", fontsize=9, color="#666")
            plt.axis("off")
            pdf.savefig(fig, bbox_inches="tight")
            plt.close(fig)

        added_to_good = 0
        for channel, name, xlsx_path, png_path in experiments:
            try:
                title, subtitle, headers, data = read_xlsx(xlsx_path)
            except Exception as e:
                print(f"  ERROR reading {xlsx_path}: {e}")
                continue
            quality = quality_idx.get(name)
            comment = make_comment(headers, data, quality, title)

            render_page(all_pdf, channel, name, title, subtitle, headers, data, png_path, comment, quality)

            include_in_good = quality == "good" or name in PRESENTATION_GOOD
            if include_in_good:
                render_page(good_pdf, channel, name, title, subtitle, headers, data, png_path, comment, quality)
                added_to_good += 1

            print(f"  [{quality or '----------'}] {channel}/{name}")

    print(f"\nWrote: {all_pdf_path}")
    print(f"Wrote: {good_pdf_path}  ({added_to_good} good experiments)")


if __name__ == "__main__":
    build_pdfs()

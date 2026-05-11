#!/usr/bin/env python3
"""Clean design-related suffixes from Excel titles and regenerate plots."""

import re
import glob
import os
import numpy as np
import openpyxl

import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
from matplotlib.ticker import ScalarFormatter

RESULTS_DIR = '/Users/ytnspybq/PycharmProjects/polar_codes_MAC/to_git_v2/results'

def strip_design_suffix(title):
    """Remove ', XXXX design' from the end of a title."""
    if title is None:
        return title
    cleaned = re.sub(r',\s*\w+\s+design\s*$', '', title, flags=re.IGNORECASE)
    return cleaned.strip()


def read_excel_data(xlsx_path):
    """Read an Excel file and return title, headers, and data rows."""
    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb.active
    title = ws.cell(row=1, column=1).value
    # Find max columns from row 4
    headers = []
    for c in range(1, 20):
        v = ws.cell(row=4, column=c).value
        if v is None:
            break
        headers.append(v)
    # Read data rows starting at row 5
    data = []
    for r in range(5, ws.max_row + 1):
        row_vals = []
        for c in range(1, len(headers) + 1):
            row_vals.append(ws.cell(row=r, column=c).value)
        if row_vals[0] is None:
            break
        data.append(row_vals)
    return title, headers, data


def make_title_lines(title, max_len=55):
    """Split a long title into 2 lines at the best comma."""
    if len(title) <= max_len:
        return title
    # Find the comma closest to the middle
    mid = len(title) // 2
    commas = [i for i, ch in enumerate(title) if ch == ',']
    if not commas:
        return title
    best = min(commas, key=lambda i: abs(i - mid))
    return title[:best+1].strip() + '\n' + title[best+1:].strip()


def detect_plot_type(headers):
    """Detect which type of plot based on headers."""
    h_set = set(headers)
    if 'SCL BLER' in h_set and 'NN BLER' in h_set and 'SC BLER' in h_set:
        return '3decoder'
    elif 'NN BLER' in h_set and 'SC BLER' in h_set:
        return '2decoder'
    elif 'BLER' in h_set:
        return 'single'
    else:
        return 'unknown'


def get_col(headers, name):
    """Get column index for a header name."""
    return headers.index(name)


def plot_single(title, headers, data, png_path):
    """Plot single decoder BLER vs N."""
    col_n = get_col(headers, 'N')
    col_bler = get_col(headers, 'BLER')
    col_ru = get_col(headers, 'Ru')
    col_rv = get_col(headers, 'Rv')

    ns = [row[col_n] for row in data]
    blers = [row[col_bler] for row in data]
    rus = [row[col_ru] for row in data]
    rvs = [row[col_rv] for row in data]

    fig, ax = plt.subplots(figsize=(9, 5.5))

    # Separate zero and nonzero
    x_vals = list(range(len(ns)))
    nonzero_x, nonzero_b = [], []
    zero_x = []
    for i, b in enumerate(blers):
        if b is not None and b > 0:
            nonzero_x.append(i)
            nonzero_b.append(b)
        else:
            zero_x.append(i)

    if nonzero_b:
        # Determine label from title
        if 'SCL' in title:
            import re as _re
            m = _re.search(r'SCL\s*\(L=(\d+)\)', title)
            lbl = f'SCL (L={m.group(1)})' if m else 'SCL'
        else:
            lbl = 'SC (L=1)'
        ax.semilogy(nonzero_x, nonzero_b, 'o-', color='#2166ac', linewidth=2, markersize=8, label=lbl)
    if zero_x:
        min_y = min(nonzero_b) if nonzero_b else 1e-4
        floor_y = max(1e-5, min_y * 0.3)
        ax.semilogy(zero_x, [floor_y]*len(zero_x), 'v', color='#2166ac', markersize=10)

    # Rate annotations
    for i in range(len(ns)):
        rate = rus[i] + rvs[i] if rus[i] is not None and rvs[i] is not None else None
        if rate is not None:
            b = blers[i] if blers[i] and blers[i] > 0 else (floor_y if zero_x else 1e-4)
            ax.annotate(f'R={rate:.3f}', (i, b), textcoords='offset points',
                       xytext=(8, 8), fontsize=7, color='gray')

    ax.set_xticks(x_vals)
    ax.set_xticklabels([str(int(n)) for n in ns])
    ax.set_xlabel('Block length N')
    ax.set_ylabel('BLER')

    # Determine ylim
    if nonzero_b:
        min_bler = min(nonzero_b)
        ax.set_ylim(bottom=1e-5 if min_bler < 1e-3 else 1e-4)

    ax.set_title(make_title_lines(title))
    ax.grid(True, which='both', alpha=0.3)
    ax.legend(fontsize=11)
    fig.tight_layout()
    fig.savefig(png_path, dpi=150)
    plt.close(fig)


def plot_2decoder(title, headers, data, png_path):
    """Plot NN vs SC comparison."""
    col_n = get_col(headers, 'N')
    col_nn = get_col(headers, 'NN BLER')
    col_sc = get_col(headers, 'SC BLER')

    # Find Ru/Rv columns
    col_ru = get_col(headers, 'Ru')
    col_rv = get_col(headers, 'Rv')

    # Find ratio column
    ratio_name = 'NN/SC' if 'NN/SC' in headers else 'Ratio'
    col_ratio = get_col(headers, ratio_name)

    ns = [row[col_n] for row in data]
    nn_blers = [row[col_nn] for row in data]
    sc_blers = [row[col_sc] for row in data]
    rus = [row[col_ru] for row in data]
    rvs = [row[col_rv] for row in data]
    ratios = [row[col_ratio] for row in data]

    fig, ax = plt.subplots(figsize=(9, 5.5))
    log2_ns = [np.log2(n) for n in ns]

    # SC line
    sc_nz_x, sc_nz_b = [], []
    sc_z_x = []
    for i, b in enumerate(sc_blers):
        if b is not None and b > 0:
            sc_nz_x.append(log2_ns[i])
            sc_nz_b.append(b)
        else:
            sc_z_x.append(log2_ns[i])

    # NN line
    nn_nz_x, nn_nz_b = [], []
    nn_z_x = []
    for i, b in enumerate(nn_blers):
        if b is not None and b > 0:
            nn_nz_x.append(log2_ns[i])
            nn_nz_b.append(b)
        else:
            nn_z_x.append(log2_ns[i])

    all_nz = (sc_nz_b or [1]) + (nn_nz_b or [1])
    min_bler = min(all_nz) if all_nz else 1e-4
    floor_y = 1e-5 if min_bler < 1e-3 else 3e-5

    if sc_nz_x:
        ax.semilogy(sc_nz_x, sc_nz_b, 's--', color='#b2182b', linewidth=2, markersize=8, label='SC (L=1)')
    if sc_z_x:
        ax.semilogy(sc_z_x, [floor_y]*len(sc_z_x), 'v', color='#b2182b', markersize=10)

    if nn_nz_x:
        ax.semilogy(nn_nz_x, nn_nz_b, 'o-', color='#2166ac', linewidth=2, markersize=8, label='Neural SC')
    if nn_z_x:
        ax.semilogy(nn_z_x, [floor_y]*len(nn_z_x), 'v', color='#2166ac', markersize=10)

    # Rate annotations (on SC line)
    for i in range(len(ns)):
        ru = rus[i] if rus[i] is not None else 0
        rv = rvs[i] if rvs[i] is not None else 0
        rate = ru + rv
        b = sc_blers[i] if sc_blers[i] and sc_blers[i] > 0 else floor_y
        ax.annotate(f'R={rate:.2f}', (log2_ns[i], b), textcoords='offset points',
                   xytext=(8, 8), fontsize=7, color='gray')

    # Ratio annotations (below NN points)
    for i in range(len(ns)):
        r = ratios[i]
        if r is not None:
            b = nn_blers[i] if nn_blers[i] and nn_blers[i] > 0 else floor_y
            color = '#2ca02c' if r < 1 else 'gray'
            ax.annotate(f'{r:.2f}', (log2_ns[i], b), textcoords='offset points',
                       xytext=(0, -14), fontsize=7, color=color, ha='center')

    ax.set_xlabel('Block length N')
    ax.set_ylabel('BLER')
    ax.xaxis.set_major_formatter(ScalarFormatter())
    ax.set_xticks(log2_ns)
    ax.set_xticklabels([str(int(n)) for n in ns])

    ax.set_ylim(bottom=1e-5 if min_bler < 1e-3 else 1e-4)
    ax.set_title(make_title_lines(title))
    ax.grid(True, which='both', alpha=0.3)
    ax.legend(fontsize=10)
    fig.tight_layout()
    fig.savefig(png_path, dpi=150)
    plt.close(fig)


def plot_3decoder(title, headers, data, png_path):
    """Plot NN vs SC vs SCL comparison."""
    col_n = get_col(headers, 'N')
    col_nn = get_col(headers, 'NN BLER')
    col_sc = get_col(headers, 'SC BLER')
    col_scl = get_col(headers, 'SCL BLER')
    col_ru = get_col(headers, 'Ru')
    col_rv = get_col(headers, 'Rv')
    col_ratio = get_col(headers, 'NN/SC')

    ns = [row[col_n] for row in data]
    nn_blers = [row[col_nn] for row in data]
    sc_blers = [row[col_sc] for row in data]
    scl_blers = [row[col_scl] for row in data]
    rus = [row[col_ru] for row in data]
    rvs = [row[col_rv] for row in data]
    ratios = [row[col_ratio] for row in data]

    fig, ax = plt.subplots(figsize=(9, 5.5))
    log2_ns = [np.log2(n) for n in ns]

    all_blers = [b for b in nn_blers + sc_blers + scl_blers if b is not None and b > 0]
    min_bler = min(all_blers) if all_blers else 1e-4
    floor_y = 1e-5 if min_bler < 1e-3 else 3e-6

    # Plot each decoder
    for blers, style, color, label in [
        (sc_blers, 's--', '#b2182b', 'SC (L=1)'),
        (scl_blers, '^-', '#2ca02c', 'SCL (L=32)'),
        (nn_blers, 'o-', '#2166ac', 'Neural SC'),
    ]:
        nz_x, nz_b, z_x = [], [], []
        for i, b in enumerate(blers):
            if b is not None and b > 0:
                nz_x.append(log2_ns[i])
                nz_b.append(b)
            else:
                z_x.append(log2_ns[i])
        if nz_x:
            ax.semilogy(nz_x, nz_b, style, color=color, linewidth=2, markersize=8, label=label)
        if z_x:
            ax.semilogy(z_x, [floor_y]*len(z_x), 'v', color=color, markersize=10)

    # Rate annotations on SC
    for i in range(len(ns)):
        ru = rus[i] if rus[i] is not None else 0
        rv = rvs[i] if rvs[i] is not None else 0
        rate = ru + rv
        b = sc_blers[i] if sc_blers[i] and sc_blers[i] > 0 else floor_y
        ax.annotate(f'R={rate:.2f}', (log2_ns[i], b), textcoords='offset points',
                   xytext=(8, 8), fontsize=7, color='gray')

    # Ratio annotations below NN
    for i in range(len(ns)):
        r = ratios[i]
        if r is not None:
            b = nn_blers[i] if nn_blers[i] and nn_blers[i] > 0 else floor_y
            color = '#2ca02c' if r < 1 else 'gray'
            ax.annotate(f'{r:.2f}', (log2_ns[i], b), textcoords='offset points',
                       xytext=(0, -14), fontsize=7, color=color, ha='center')

    ax.set_xlabel('Block length N')
    ax.set_ylabel('BLER')
    ax.xaxis.set_major_formatter(ScalarFormatter())
    ax.set_xticks(log2_ns)
    ax.set_xticklabels([str(int(n)) for n in ns])

    ax.set_ylim(bottom=1e-5 if min_bler < 1e-3 else 1e-4)
    ax.set_title(make_title_lines(title))
    ax.grid(True, which='both', alpha=0.3)
    ax.legend(fontsize=10)
    fig.tight_layout()
    fig.savefig(png_path, dpi=150)
    plt.close(fig)


def process_all():
    xlsx_files = sorted(glob.glob(os.path.join(RESULTS_DIR, '**', '*.xlsx'), recursive=True))

    for xlsx_path in xlsx_files:
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active
        old_title = ws.cell(row=1, column=1).value
        new_title = strip_design_suffix(old_title)

        title_changed = (old_title != new_title)

        if title_changed:
            ws.cell(row=1, column=1).value = new_title
            wb.save(xlsx_path)
            rel = xlsx_path.split('/results/')[1]
            print(f'[XLSX] Cleaned title: {rel}')
            print(f'       "{old_title}" -> "{new_title}"')

        # Find matching PNG in same directory
        xlsx_dir = os.path.dirname(xlsx_path)
        png_files = glob.glob(os.path.join(xlsx_dir, '*.png'))
        # Filter out no256 files
        png_files = [p for p in png_files if 'no256' not in os.path.basename(p)]

        if not png_files:
            continue

        # Re-read data (possibly with updated title)
        title, headers, data = read_excel_data(xlsx_path)
        plot_type = detect_plot_type(headers)

        for png_path in png_files:
            rel_png = png_path.split('/results/')[1]
            if plot_type == 'single':
                plot_single(title, headers, data, png_path)
                print(f'[PNG]  Regenerated (single): {rel_png}')
            elif plot_type == '2decoder':
                plot_2decoder(title, headers, data, png_path)
                print(f'[PNG]  Regenerated (2-decoder): {rel_png}')
            elif plot_type == '3decoder':
                plot_3decoder(title, headers, data, png_path)
                print(f'[PNG]  Regenerated (3-decoder): {rel_png}')
            else:
                print(f'[PNG]  SKIPPED (unknown type): {rel_png}')


if __name__ == '__main__':
    process_all()
    print('\nDone!')

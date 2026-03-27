#!/usr/bin/env python
"""
extend_classB_Ru39_10K.py — Extend GMAC 3dB Class B Ru39/Rv39 SC to 10K CW
for N=128,256,512,1024 using the optimized v3 decoder.
"""

import os
import sys
import time
import numpy as np
from datetime import datetime

# Use v3 optimized decoder
sys.path.insert(0, '/Users/ytnspybq/PycharmProjects/polar_codes_MAC/to_git_v3')
from polar.channels import GaussianMAC
from polar.eval import MACEval
from polar.design import make_path

# v2 design loader
sys.path.insert(0, '/Users/ytnspybq/PycharmProjects/polar_codes_MAC/to_git_v2')
from polar.design_mc import load_design, _select_info_frozen

import openpyxl
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt

DESIGNS_DIR = '/Users/ytnspybq/PycharmProjects/polar_codes_MAC/to_git_v2/designs'
RESULTS_DIR = '/Users/ytnspybq/PycharmProjects/polar_codes_MAC/to_git_v2/results/gmac_snr3dB/gmac_snr3dB_classB_Ru39_Rv39_SC'
XLSX_PATH = os.path.join(RESULTS_DIR, 'classB_Ru39_Rv39_snr3dB_SC.xlsx')
PLOT_PATH = os.path.join(RESULTS_DIR, 'bler_vs_N_classB_Ru39_Rv39_snr3dB_SC.png')
AUDIT_PATH = '/Users/ytnspybq/PycharmProjects/polar_codes_MAC/to_git_v2/results/RESULTS_AUDIT.xlsx'


def log(msg=""):
    print(f"[{datetime.now().strftime('%H:%M:%S')}] {msg}", flush=True)


def read_existing_xlsx(path):
    """Read existing xlsx results."""
    wb = openpyxl.load_workbook(path)
    ws = wb.active
    data = {}
    for row in ws.iter_rows(min_row=5, max_row=12, values_only=True):
        if row[0] and isinstance(row[0], (int, float)) and row[0] > 1:
            N = int(row[0])
            ku, kv = int(row[1]), int(row[2])
            cw = int(row[7])
            bler = float(row[6])
            data[N] = {
                "ku": ku, "kv": kv, "cw": cw,
                "err": round(bler * cw)
            }
    return data


def main():
    sigma2 = 10.0 ** (-3.0 / 10.0)
    channel = GaussianMAC(sigma2=sigma2)
    I_ZX, I_ZY_X, I_ZXY = channel.capacity()
    cap_line = f"Capacity: I(Z;X)={I_ZX:.4f}, I(Z;Y|X)={I_ZY_X:.4f}, I(Z;X,Y)={I_ZXY:.4f}"
    log(f"GMAC 3dB: {cap_line}")

    # Read existing results
    existing = read_existing_xlsx(XLSX_PATH)
    log(f"Existing data: {[(N, existing[N]['cw']) for N in sorted(existing.keys())]}")

    N_LIST = [8, 16, 32, 64, 128, 256, 512, 1024]
    TARGET_CW = {128: 10000, 256: 10000, 512: 10000, 1024: 10000}

    # Set up evaluator — use single worker to keep CPU moderate
    evaluator = MACEval(channel, log_domain=True, n_workers=1,
                        rng=np.random.default_rng(424242),
                        decoder_type='sc', backend='auto')

    results = {}
    for N in N_LIST:
        ex = existing[N]
        ku, kv = ex["ku"], ex["kv"]
        target = TARGET_CW.get(N, ex["cw"])
        new_needed = max(0, target - ex["cw"])

        if new_needed <= 0:
            log(f"  N={N:>5d}: already have {ex['cw']} CW -- SKIP")
            results[N] = {"err": ex["err"], "cw": ex["cw"],
                          "ku": ku, "kv": kv}
            continue

        log(f"  N={N:>5d}: running {new_needed} new CW "
            f"(have {ex['cw']}, target {target})")

        n = int(np.log2(N))
        path_i = N // 2  # Class B

        # Load MC design
        design_file = os.path.join(DESIGNS_DIR, f"gmac_B_n{n}_snr3dB.npz")
        su, sv, _, _, pi = load_design(design_file)
        Au, fu = _select_info_frozen(N, su, ku)
        Av, fv = _select_info_frozen(N, sv, kv)
        b = make_path(N, path_i=path_i)

        t0 = time.time()
        # Use different seed from existing
        evaluator.rng = np.random.default_rng(777777 + N * 1000)
        ber_u, ber_v, bler_new = evaluator.run(
            N, b, Au, Av, fu, fv,
            n_codewords=new_needed,
            batch_size=50 if N >= 512 else 100,
            verbose=True)

        new_err = round(bler_new * new_needed)
        total_err = ex["err"] + new_err
        total_cw = ex["cw"] + new_needed
        bler = total_err / total_cw
        elapsed = time.time() - t0
        log(f"  N={N:>5d}: BLER={bler:.6f} ({total_err}/{total_cw}) "
            f"in {elapsed:.1f}s")
        results[N] = {"err": total_err, "cw": total_cw, "ku": ku, "kv": kv}

    # Write xlsx
    title = "Gaussian MAC \u2014 Class B, (Ru=0.39, Rv=0.39), SNR=3dB, SC (L=1)"
    subtitle = ("Symmetric capacity: Rx=Ry \u2264 0.5548 bits/use | "
                "Sum capacity: 1.1095 bits/use | Path: 0^(N/2) 1^N 0^(N/2)")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append([title] + [None] * 7)
    ws.append([subtitle] + [None] * 7)
    ws.append([None] * 8)
    ws.append(['N', 'ku', 'kv', 'Ru', 'Rv', 'Sum Rate', 'BLER', 'Codewords'])
    for N in N_LIST:
        r = results[N]
        ku, kv = r["ku"], r["kv"]
        Ru, Rv = ku / N, kv / N
        bler = r["err"] / r["cw"] if r["cw"] > 0 else 0
        ws.append([N, ku, kv, round(Ru, 4), round(Rv, 4),
                   round(Ru + Rv, 4), round(bler, 6), r["cw"]])
    ws.append([None] * 8)
    ws.append([cap_line] + [None] * 7)
    wb.save(XLSX_PATH)
    log(f"Saved xlsx: {XLSX_PATH}")

    # Generate plot
    Ns = []
    blers = []
    for N in N_LIST:
        r = results[N]
        bler = r["err"] / r["cw"] if r["cw"] > 0 else None
        if bler is not None and bler > 0:
            Ns.append(N)
            blers.append(bler)

    fig, ax = plt.subplots(figsize=(8, 5))
    ax.semilogy(range(len(Ns)), blers, 'bo-', linewidth=2, markersize=8)
    ax.set_xticks(range(len(Ns)))
    ax.set_xticklabels([str(n) for n in Ns])
    ax.set_xlabel('Block Length N')
    ax.set_ylabel('BLER')
    ax.set_title('GMAC 3dB, Class B, Ru=Rv=0.39, SC')
    ax.grid(True, alpha=0.3)
    for i, (n, b) in enumerate(zip(Ns, blers)):
        ax.annotate(f'{b:.4f}', (i, b), textcoords="offset points",
                    xytext=(0, 10), ha='center', fontsize=8)
    plt.tight_layout()
    plt.savefig(PLOT_PATH, dpi=150)
    log(f"Saved plot: {PLOT_PATH}")

    # Update RESULTS_AUDIT.xlsx
    import pandas as pd
    df = pd.read_excel(AUDIT_PATH)

    for N in N_LIST:
        r = results[N]
        ku, kv = r["ku"], r["kv"]
        bler = r["err"] / r["cw"] if r["cw"] > 0 else 0
        errors = r["err"]
        cw = r["cw"]

        # Determine confidence/status
        if errors >= 100:
            conf, stat = "HIGH", "GOOD"
        elif errors >= 30:
            conf, stat = "HIGH", "GOOD"
        elif errors >= 10:
            conf, stat = "MEDIUM", "OK"
        elif errors > 0:
            conf, stat = "LOW", "NEEDS_MORE_CW"
        else:
            conf, stat = "ZERO", "NEEDS_MORE_CW"

        Ru, Rv = round(ku / N, 4), round(kv / N, 4)

        # Find matching row — use abs tolerance for float matching
        mask = ((df['Channel'] == 'GMAC') & (df['SNR'] == '3dB') &
                (df['Class'] == 'B') & (df['Decoder'] == 'SC') &
                (df['N'] == N) &
                ((df['Ru'] - Ru).abs() < 0.005) &
                ((df['Rv'] - Rv).abs() < 0.005))
        idx = df.index[mask]

        if len(idx) > 0:
            i = idx[0]
            df.at[i, 'BLER'] = round(bler, 6)
            df.at[i, 'Errors'] = errors
            df.at[i, 'Codewords'] = cw
            df.at[i, 'Confidence'] = conf
            df.at[i, 'Status'] = stat
            if cw > existing.get(N, {}).get('cw', 0):
                df.at[i, 'Notes'] = f'Extended to {cw} CW'
            log(f"  AUDIT: Updated row {i} for N={N}")
        else:
            log(f"  AUDIT: No matching row for N={N}, Ru={Ru} -- skipping")

    df.to_excel(AUDIT_PATH, index=False)
    log(f"Saved audit: {AUDIT_PATH}")

    log("\nFinal results:")
    for N in N_LIST:
        r = results[N]
        bler = r["err"] / r["cw"] if r["cw"] > 0 else 0
        log(f"  N={N:>5d}: BLER={bler:.6f} ({r['err']}/{r['cw']})")

    log("\nDONE!")


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        log("\n*** Interrupted ***")
        sys.exit(1)
